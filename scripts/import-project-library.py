#!/usr/bin/env python3
"""Controlled Mettelo Project Library importer.

Default mode is DRY RUN. The workbook is the editorial source of truth; Supabase
is the runtime source of truth. Apply mode requires SUPABASE_URL,
SUPABASE_SERVICE_ROLE_KEY, and an explicit production-write authorization phrase.
Reconciliation is deterministic by canonical Project ID, slug, then exact normalised
title. Ambiguous matches are never updated automatically.
"""
from __future__ import annotations

import argparse
import hashlib
import json
import os
import re
import unicodedata
import urllib.error
import urllib.parse
import urllib.request
from dataclasses import dataclass
from pathlib import Path
from typing import Any, Iterable

try:
    from openpyxl import load_workbook
except ImportError as exc:
    raise SystemExit("openpyxl is required: python -m pip install openpyxl") from exc

PROJECT_SHEET = "03_PROJECT_LIBRARY"
HEADER_ROW = 4
EXPECTED_PROJECT_COUNT = 300
WRITE_AUTHORIZATION_ENV = "PROJECT_LIBRARY_WRITE_AUTHORIZATION"
WRITE_AUTHORIZATION_PHRASE = "I AUTHORIZE THE PRODUCTION PROJECT LIBRARY APPLY"
REQUIRED_HEADERS = {
    "Project ID", "Project Title", "Industry / Domain", "Dataset", "Source",
    "Data Link", "Licence / Reuse", "Data Reality", "Stakeholder",
    "Problem Statement (200+ words)", "Use Case (200+ words)",
    "Decision to Support", "Project Objective", "Specific Deliverables",
    "Success Criteria", "Technical Skills", "Professional Skills",
    "Team / Roles", "Tools", "Methods", "Difficulty", "Duration",
    "Weekly Commitment", "Evidence / Proof", "Constraints / Trade-offs",
    "Explicit Assumptions", "Out of Scope", "Role Responsibilities",
    "Acceptance / Quality Checks", "Risks / Responsible Use",
    "Stakeholder Handover", "Preservation Class", "Mettelo May Store Copy?",
    "Mettelo May Redistribute?", "Attribution Required?", "Legal / Provenance Note",
    "Preservation Mode", "Exact Data to Download / Preserve", "Member Dataset Scope",
}


def text(value: Any) -> str:
    return "" if value is None else str(value).strip()


def slugify(value: str) -> str:
    raw = unicodedata.normalize("NFKD", value).encode("ascii", "ignore").decode().lower()
    raw = re.sub(r"[^a-z0-9]+", "-", raw).strip("-")
    return raw[:120] or "project"


def canonical_slug(title: str, project_id: str) -> str:
    suffix = project_id.lower()
    base = slugify(title)
    max_base = max(1, 120 - len(suffix) - 1)
    return f"{base[:max_base].rstrip('-')}-{suffix}"


def normal_title(value: str) -> str:
    return re.sub(r"\s+", " ", value).strip().casefold()


def normal_difficulty(value: Any) -> str:
    raw = text(value).casefold()
    if "advanced" in raw:
        return "advanced"
    if "intermediate" in raw or "applied" in raw:
        return "intermediate"
    return "entry"


def normal_source_type(source: Any, url: Any) -> str:
    raw = f"{text(source)} {text(url)}".casefold()
    if "kaggle" in raw:
        return "kaggle"
    if "huggingface" in raw or "hugging face" in raw:
        return "hugging_face"
    if "github" in raw:
        return "github"
    if "api" in raw or "open-meteo" in raw:
        return "api"
    if any(token in raw for token in (
        "gov.uk", "data.gov", "statistics.gov", "ons.gov", "nhs", "worldbank",
        "world bank", "opendata", "open data", "archive.ics.uci", "uci machine learning",
    )):
        return "public_portal"
    return "external_website"


def list_items(value: Any) -> list[str]:
    raw = text(value)
    if not raw:
        return []
    lines = [x.strip() for x in re.split(r"\r?\n", raw) if x.strip()]
    if len(lines) == 1 and ";" in raw:
        lines = [x.strip() for x in raw.split(";") if x.strip()]
    out: list[str] = []
    for line in lines:
        cleaned = re.sub(r"^\s*(?:[-*•]|\d{1,3}[.)])\s*", "", line).strip()
        if cleaned:
            out.append(cleaned)
    return out


def csv_items(value: Any) -> list[str]:
    raw = text(value)
    if not raw:
        return []
    return [x.strip() for x in re.split(r"\r?\n|,|;", raw) if x.strip()]


def duration_weeks(value: Any) -> int | None:
    nums = [int(x) for x in re.findall(r"\d+", text(value))]
    if not nums:
        return None
    return max(nums) if len(nums) > 1 else nums[0]


def role_specs(value: Any) -> tuple[list[str], int | None]:
    """Parse every approved workbook role grammar and return declared team size."""
    raw = text(value)
    if not raw:
        return [], None
    declared: int | None = None
    prefix = re.match(r"^\s*(\d+)\s+people\s+(?:—|–|-)\s+(.+)$", raw, flags=re.I | re.S)
    if not prefix:
        prefix = re.match(r"^\s*(\d+)\s*[- ]person\s+team\s*:\s*(.+)$", raw, flags=re.I | re.S)
    if prefix:
        declared = int(prefix.group(1))
        raw = prefix.group(2).strip()
    elif re.match(r"^\s*solo\s+project\b", raw, flags=re.I):
        declared = 1

    if "\n" in raw:
        specs = list_items(raw)
    elif ";" in raw:
        specs = [x.strip() for x in raw.split(";") if x.strip()]
    elif "," in raw and (declared is not None or re.search(r"(?:^|,\s*)\d+\s+", raw)):
        specs = [x.strip() for x in raw.split(",") if x.strip()]
    elif " — " in raw or " – " in raw:
        specs = [x.strip() for x in re.split(r"(?<=\.)\s+(?=[A-Z][^.]+?\s+(?:—|–))", raw) if x.strip()]
    else:
        specs = [raw]
    return specs, declared


def clean_role_label(value: str) -> str:
    value = re.sub(r"^\s*(?:[-*•]|\d{1,3}[.)])\s*", "", value).strip()
    value = re.sub(r"^\d+\s+", "", value).strip()
    return value.rstrip(".").strip()


def parse_roles(value: Any, responsibilities: Any) -> tuple[list[dict[str, Any]], int | None]:
    specs, declared = role_specs(value)
    parsed_specs: list[tuple[str, int, str | None]] = []
    for spec in specs:
        cleaned = re.sub(r"^\s*(?:[-*•]|\d{1,3}[.)])\s*", "", spec).strip()
        if not cleaned or normal_title(cleaned).startswith(("all members", "cross-review rule")):
            continue
        match = re.match(r"^(\d+)\s+(.+)$", cleaned)
        capacity = int(match.group(1)) if match else 1
        body = match.group(2).strip() if match else cleaned
        parts = re.split(r"\s+(?:—|–)\s+|:\s*", body, maxsplit=1)
        name = clean_role_label(parts[0])
        inline = parts[1].strip() if len(parts) == 2 and parts[1].strip() else None
        if name and not normal_title(name).startswith(("all members", "cross-review rule")):
            parsed_specs.append((name, capacity, inline))

    role_names = {normal_title(name) for name, _, _ in parsed_specs}
    resp_map: dict[str, list[str]] = {}
    shared: list[str] = []
    for line in list_items(responsibilities):
        if normal_title(line).startswith(("all members", "cross-review rule")):
            continue
        parts = re.split(r"\s+(?:—|–)\s+|:\s*", line, maxsplit=1)
        if len(parts) == 2:
            label = normal_title(clean_role_label(parts[0]))
            body = parts[1].strip()
            if label in role_names:
                resp_map.setdefault(label, []).append(body)
            elif body:
                shared.append(body)
        else:
            shared.append(line.strip())

    roles: list[dict[str, Any]] = []
    for name, capacity, inline in parsed_specs:
        direct = list(resp_map.get(normal_title(name), []))
        if not direct and inline:
            direct = [inline]
        if not direct and shared:
            direct = list(shared)
        roles.append({"name": name, "capacity": capacity, "responsibilities": direct})
    return roles, declared


def governance_status(value: Any) -> str:
    v = text(value).upper()
    if v == "GREEN":
        return "green"
    if v == "AMBER":
        return "amber"
    if v == "RED":
        return "red"
    return "verification_required"


def yes(value: Any) -> bool:
    return text(value).casefold().startswith("yes")


def short_title(value: str, limit: int = 180) -> str:
    value = re.sub(r"\s+", " ", value).strip()
    return value if len(value) <= limit else value[: limit - 1].rstrip() + "…"


def same_payload(row: dict[str, Any] | None, payload: dict[str, Any]) -> bool:
    if not row:
        return False
    return all(row.get(key) == value for key, value in payload.items())


@dataclass
class Api:
    base: str
    key: str

    def request(self, method: str, path: str, payload: Any | None = None, prefer: str | None = None) -> Any:
        body = None if payload is None else json.dumps(payload).encode()
        headers = {"apikey": self.key, "Authorization": f"Bearer {self.key}", "Accept": "application/json"}
        if body is not None:
            headers["Content-Type"] = "application/json"
        if prefer:
            headers["Prefer"] = prefer
        req = urllib.request.Request(
            self.base.rstrip("/") + "/rest/v1/" + path.lstrip("/"),
            data=body,
            method=method,
            headers=headers,
        )
        try:
            with urllib.request.urlopen(req, timeout=90) as res:
                raw = res.read().decode()
                return json.loads(raw) if raw else None
        except urllib.error.HTTPError as exc:
            detail = exc.read().decode(errors="replace")
            raise RuntimeError(f"Supabase {method} {path} failed: HTTP {exc.code}: {detail}") from exc

    def existing_projects(self) -> list[dict[str, Any]]:
        return self.request("GET", "projects?select=id,canonical_project_key,slug,title&limit=5000") or []

    def patch(self, table: str, query: str, payload: dict[str, Any]) -> None:
        self.request("PATCH", f"{table}?{query}", payload, "return=minimal")

    def insert(self, table: str, payload: dict[str, Any]) -> dict[str, Any]:
        rows = self.request("POST", table, payload, "return=representation") or []
        return rows[0] if isinstance(rows, list) and rows else rows

    def upsert(self, table: str, conflict: str, payload: dict[str, Any]) -> None:
        path = f"{table}?on_conflict={urllib.parse.quote(conflict)}"
        self.request("POST", path, payload, "resolution=merge-duplicates,return=minimal")


def existing_child(rows: list[dict[str, Any]], key_field: str, key: str, title_field: str, title: str) -> dict[str, Any] | None:
    keyed = [row for row in rows if text(row.get(key_field)) == key]
    if len(keyed) > 1:
        raise RuntimeError(f"Ambiguous existing canonical child key {key}")
    if keyed:
        return keyed[0]
    titled = [row for row in rows if normal_title(text(row.get(title_field))) == normal_title(title)]
    if len(titled) > 1:
        raise RuntimeError(f"Ambiguous legacy child title {title!r}")
    return titled[0] if titled else None


def load_records(path: Path) -> tuple[list[dict[str, Any]], list[str]]:
    wb = load_workbook(path, read_only=True, data_only=True)
    if PROJECT_SHEET not in wb.sheetnames:
        raise ValueError(f"Missing sheet {PROJECT_SHEET!r}")
    ws = wb[PROJECT_SHEET]
    headers = [text(c.value) for c in ws[HEADER_ROW]]
    missing_headers = sorted(REQUIRED_HEADERS - set(headers))
    if missing_headers:
        raise ValueError("Missing required headers: " + ", ".join(missing_headers))
    pos = {h: i for i, h in enumerate(headers) if h}
    records: list[dict[str, Any]] = []
    issues: list[str] = []
    seen: set[str] = set()
    for row_num, row in enumerate(ws.iter_rows(min_row=HEADER_ROW + 1, values_only=True), start=HEADER_ROW + 1):
        pid = text(row[pos["Project ID"]])
        if not pid:
            continue
        if pid in seen:
            issues.append(f"DUPLICATE_PROJECT_ID:{pid}:row={row_num}")
        seen.add(pid)
        get = lambda k: row[pos[k]]
        title = text(get("Project Title"))
        data_link = text(get("Data Link"))
        roles, declared_team_size = parse_roles(get("Team / Roles"), get("Role Responsibilities"))
        deliverables = list_items(get("Specific Deliverables"))
        criteria = list_items(get("Success Criteria"))
        if not deliverables:
            issues.append(f"MISSING_DELIVERABLES:{pid}")
        if not criteria:
            issues.append(f"MISSING_SUCCESS_CRITERIA:{pid}")
        if not roles:
            issues.append(f"MISSING_ROLES:{pid}")
        if roles and any(not role["responsibilities"] for role in roles):
            issues.append(f"MISSING_ROLE_RESPONSIBILITIES:{pid}")
        if not data_link.lower().startswith("https://"):
            issues.append(f"INVALID_DATA_LINK:{pid}")
        team_size = sum(int(role["capacity"]) for role in roles)
        if team_size not in {1, 3, 4, 5}:
            issues.append(f"INVALID_TEAM_SIZE:{pid}:{team_size}")
        if declared_team_size is not None and team_size != declared_team_size:
            issues.append(f"TEAM_SIZE_DECLARATION_MISMATCH:{pid}:declared={declared_team_size}:parsed={team_size}")
        records.append({
            "project_id": pid,
            "title": title,
            "slug": canonical_slug(title, pid),
            "domain": text(get("Industry / Domain")),
            "dataset": text(get("Dataset")),
            "source": text(get("Source")),
            "source_type": normal_source_type(get("Source"), data_link),
            "data_link": data_link,
            "licence": text(get("Licence / Reuse")),
            "data_reality": text(get("Data Reality")),
            "stakeholder": text(get("Stakeholder")),
            "problem_statement": text(get("Problem Statement (200+ words)")),
            "use_case": text(get("Use Case (200+ words)")),
            "decision_to_support": text(get("Decision to Support")),
            "objective": text(get("Project Objective")),
            "deliverables": deliverables,
            "success_criteria": criteria,
            "technical_skills": csv_items(get("Technical Skills")),
            "professional_skills": csv_items(get("Professional Skills")),
            "roles": roles,
            "tools": csv_items(get("Tools")),
            "methods": csv_items(get("Methods")),
            "difficulty": normal_difficulty(get("Difficulty")),
            "duration": text(get("Duration")),
            "duration_weeks": duration_weeks(get("Duration")),
            "weekly_commitment": text(get("Weekly Commitment")),
            "team_size": team_size,
            "evidence": list_items(get("Evidence / Proof")),
            "constraints": list_items(get("Constraints / Trade-offs")),
            "assumptions": list_items(get("Explicit Assumptions")),
            "out_of_scope": list_items(get("Out of Scope")),
            "acceptance_checks": list_items(get("Acceptance / Quality Checks")),
            "responsible_use": list_items(get("Risks / Responsible Use")),
            "handover": text(get("Stakeholder Handover")),
            "preservation_class": text(get("Preservation Class")),
            "may_store": yes(get("Mettelo May Store Copy?")),
            "may_redistribute": yes(get("Mettelo May Redistribute?")),
            "attribution": text(get("Attribution Required?")),
            "legal_note": text(get("Legal / Provenance Note")),
            "preservation_mode": text(get("Preservation Mode")),
            "preserve_scope": text(get("Exact Data to Download / Preserve")),
            "member_dataset_scope": text(get("Member Dataset Scope")),
        })
    if len(records) != EXPECTED_PROJECT_COUNT:
        issues.append(f"WORKBOOK_PROJECT_COUNT:expected={EXPECTED_PROJECT_COUNT}:actual={len(records)}")
    return records, issues


def indexes(existing: Iterable[dict[str, Any]]):
    by_key: dict[str, list[dict[str, Any]]] = {}
    by_slug: dict[str, list[dict[str, Any]]] = {}
    by_title: dict[str, list[dict[str, Any]]] = {}
    for p in existing:
        key = text(p.get("canonical_project_key"))
        if key:
            by_key.setdefault(key, []).append(p)
        slug = text(p.get("slug"))
        if slug:
            by_slug.setdefault(slug, []).append(p)
        title = normal_title(text(p.get("title")))
        if title:
            by_title.setdefault(title, []).append(p)
    return by_key, by_slug, by_title


def match_record(record: dict[str, Any], idxs):
    by_key, by_slug, by_title = idxs
    candidates = by_key.get(record["project_id"], [])
    rule = "canonical_project_key"
    if not candidates:
        candidates = by_slug.get(record["slug"], [])
        rule = "slug"
    if not candidates:
        candidates = by_title.get(normal_title(record["title"]), [])
        rule = "exact_normalised_title"
    if len(candidates) > 1:
        return None, f"AMBIGUOUS:{rule}:{len(candidates)}"
    if len(candidates) == 1:
        return candidates[0], rule
    return None, "new"


def project_payload(r: dict[str, Any], new: bool) -> dict[str, Any]:
    payload = {
        "canonical_project_key": r["project_id"],
        "title": r["title"],
        "summary": short_title(r["objective"], 900),
        "problem_statement": r["problem_statement"],
        "difficulty_level": r["difficulty"] or None,
        "duration_weeks": r["duration_weeks"],
        "weekly_commitment": r["weekly_commitment"] or None,
        "team_size_threshold": r["team_size"],
    }
    if new:
        payload.update({
            "slug": r["slug"],
            "status": "draft",
            "visibility": "private",
            "project_type": "open",
            "applications_open": False,
        })
    return payload


def apply_record(api: Api, r: dict[str, Any], match: dict[str, Any] | None) -> tuple[str, dict[str, int]]:
    writes = {"project": 0, "brief": 0, "deliverables": 0, "success_criteria": 0, "roles": 0, "sources": 0}
    was_existing = match is not None
    if match:
        project_uuid = str(match["id"])
        payload = project_payload(r, False)
        select_fields = ",".join(["id", *payload.keys()])
        current_rows = api.request(
            "GET",
            "projects?select=" + urllib.parse.quote(select_fields, safe=",") +
            "&id=eq." + urllib.parse.quote(project_uuid) + "&limit=1",
        ) or []
        current = current_rows[0] if current_rows else None
        if not same_payload(current, payload):
            api.patch("projects", "id=eq." + urllib.parse.quote(project_uuid), payload)
            writes["project"] += 1
    else:
        inserted = api.insert("projects", project_payload(r, True))
        project_uuid = str(inserted["id"])
        writes["project"] += 1

    brief = {
        "project_id": project_uuid,
        "context": r["use_case"],
        "stakeholder": r["stakeholder"],
        "expected_outcome": r["objective"],
        "success_metrics": "\n".join(f"- {x}" for x in r["success_criteria"]),
        "primary_use_case": r["use_case"],
        "primary_objective": r["objective"],
        "supporting_objectives": [],
        "key_questions": [],
        "in_scope": [],
        "out_of_scope": r["out_of_scope"],
        "decision_to_support": r["decision_to_support"],
        "constraints_trade_offs": r["constraints"],
        "explicit_assumptions": r["assumptions"],
        "acceptance_quality_checks": r["acceptance_checks"],
        "responsible_use_risks": r["responsible_use"],
        "evidence_expectations": r["evidence"],
        "technical_skills": r["technical_skills"],
        "professional_skills": r["professional_skills"],
        "canonical_methods": r["methods"],
        "canonical_tools": r["tools"],
        "stakeholder_handover": r["handover"],
    }
    qid = urllib.parse.quote(project_uuid)
    existing_briefs = api.request("GET", f"project_problem_briefs?select=*&project_id=eq.{qid}&limit=1") or []
    if existing_briefs:
        if not same_payload(existing_briefs[0], brief):
            api.patch("project_problem_briefs", "project_id=eq." + qid, brief)
            writes["brief"] += 1
    else:
        api.upsert("project_problem_briefs", "project_id", brief)
        writes["brief"] += 1

    existing_deliverables = api.request("GET", f"project_deliverables?select=*&project_id=eq.{qid}&project_run_id=is.null") or []
    existing_criteria = api.request("GET", f"project_success_criteria?select=*&project_id=eq.{qid}") or []
    existing_roles = api.request("GET", f"project_roles?select=*&project_id=eq.{qid}") or []
    existing_sources = api.request("GET", f"project_data_sources?select=*&project_id=eq.{qid}&project_run_id=is.null") or []

    for i, item in enumerate(r["deliverables"], 1):
        key = f"{r['project_id']}:deliverable:{i:03d}"
        payload = {
            "project_id": project_uuid,
            "project_run_id": None,
            "canonical_item_key": key,
            "title": short_title(item),
            "deliverable_type": "canonical",
            "acceptance_criteria": item,
            "public_summary": item,
            "is_required": True,
            "status": "planned",
            "sort_order": i,
        }
        existing = existing_child(existing_deliverables, "canonical_item_key", key, "title", payload["title"])
        if existing:
            if not same_payload(existing, payload):
                api.patch("project_deliverables", "id=eq." + urllib.parse.quote(str(existing["id"])), payload)
                writes["deliverables"] += 1
        else:
            api.upsert("project_deliverables", "project_id,canonical_item_key", payload)
            writes["deliverables"] += 1

    for i, item in enumerate(r["success_criteria"], 1):
        key = f"{r['project_id']}:criterion:{i:03d}"
        payload = {
            "project_id": project_uuid,
            "canonical_item_key": key,
            "title": short_title(item),
            "description": item,
            "is_required": True,
            "visibility": "public",
            "sort_order": i,
        }
        existing = existing_child(existing_criteria, "canonical_item_key", key, "title", payload["title"])
        if existing:
            if not same_payload(existing, payload):
                api.patch("project_success_criteria", "id=eq." + urllib.parse.quote(str(existing["id"])), payload)
                writes["success_criteria"] += 1
        else:
            api.upsert("project_success_criteria", "project_id,canonical_item_key", payload)
            writes["success_criteria"] += 1

    role_skills = list(dict.fromkeys(r["technical_skills"] + r["professional_skills"]))
    for i, role in enumerate(r["roles"], 1):
        key = f"{r['project_id']}:role:{i:02d}"
        payload = {
            "project_id": project_uuid,
            "canonical_role_key": key,
            "title": role["name"],
            "description": role["responsibilities"][0] if role["responsibilities"] else None,
            "skills": role_skills,
            "openings": role["capacity"],
            "responsibilities": role["responsibilities"],
            "recommended_skills": role_skills,
            "weekly_commitment": r["weekly_commitment"] or None,
        }
        existing = existing_child(existing_roles, "canonical_role_key", key, "title", role["name"])
        if existing:
            if not same_payload(existing, payload):
                api.patch("project_roles", "id=eq." + urllib.parse.quote(str(existing["id"])), payload)
                writes["roles"] += 1
        else:
            payload["role_status"] = "open"
            api.upsert("project_roles", "project_id,canonical_role_key", payload)
            writes["roles"] += 1

    gov = governance_status(r["preservation_class"])
    source_key = f"{r['project_id']}:source:01"
    source_payload = {
        "project_id": project_uuid,
        "project_run_id": None,
        "canonical_source_key": source_key,
        "name": r["dataset"],
        "description": r["member_dataset_scope"] or r["data_reality"],
        "source_type": r["source_type"],
        "external_url": r["data_link"],
        "provider_name": r["source"] or None,
        "licence_name": r["licence"] or None,
        "required_subset": r["member_dataset_scope"] or r["preserve_scope"] or None,
        "provenance": r["legal_note"] or None,
        "sensitivity": "public",
        "publish_policy": "permitted" if gov == "green" else "restricted",
        "governance_status": gov,
        "retention_policy": "permitted" if r["may_store"] else "not_permitted",
        "internal_storage_policy": "permitted" if r["may_store"] else "not_permitted",
    }
    existing = existing_child(existing_sources, "canonical_source_key", source_key, "name", r["dataset"])
    if existing:
        if not same_payload(existing, source_payload):
            api.patch("project_data_sources", "id=eq." + urllib.parse.quote(str(existing["id"])), source_payload)
            writes["sources"] += 1
    else:
        api.upsert("project_data_sources", "project_id,canonical_source_key", source_payload)
        writes["sources"] += 1

    if not was_existing:
        action = "created"
    elif sum(writes.values()) == 0:
        action = "unchanged"
    else:
        action = "updated"
    return action, writes


def main() -> int:
    parser = argparse.ArgumentParser()
    parser.add_argument("workbook", type=Path)
    parser.add_argument("--apply", action="store_true", help="Apply to Supabase. Default is dry-run.")
    parser.add_argument("--report", type=Path, default=Path("project-library-import-report.json"))
    args = parser.parse_args()

    raw = args.workbook.read_bytes()
    records, issues = load_records(args.workbook)
    duplicate_ids = [x for x in issues if x.startswith("DUPLICATE_PROJECT_ID")]
    required_failures = [x for x in issues if x.startswith((
        "MISSING_", "INVALID_TEAM_SIZE", "INVALID_DATA_LINK",
        "TEAM_SIZE_DECLARATION_MISMATCH", "WORKBOOK_PROJECT_COUNT",
    ))]
    report: dict[str, Any] = {
        "source": str(args.workbook),
        "source_sha256": hashlib.sha256(raw).hexdigest(),
        "workbook_projects": len(records),
        "unique_project_ids": len({r["project_id"] for r in records}),
        "duplicate_project_ids": duplicate_ids,
        "validation_issues": issues,
        "projects_with_complete_role_responsibilities": sum(
            1 for r in records if r["roles"] and all(role["responsibilities"] for role in r["roles"])
        ),
        "team_size_distribution": {},
        "preservation_distribution": {},
        "backend_before": None,
        "matched": 0,
        "to_update": 0,
        "to_create": 0,
        "ambiguous_matches": [],
        "unmatched_existing": [],
        "apply": bool(args.apply),
    }
    for r in records:
        team_key = str(r["team_size"])
        report["team_size_distribution"][team_key] = report["team_size_distribution"].get(team_key, 0) + 1
        preservation = r["preservation_class"]
        report["preservation_distribution"][preservation] = report["preservation_distribution"].get(preservation, 0) + 1

    url = os.getenv("SUPABASE_URL")
    key = os.getenv("SUPABASE_SERVICE_ROLE_KEY")
    api = Api(url, key) if url and key else None
    existing: list[dict[str, Any]] = []
    matches: dict[str, tuple[dict[str, Any] | None, str]] = {}
    if api:
        existing = api.existing_projects()
        report["backend_before"] = len(existing)
        idxs = indexes(existing)
        matched_ids: set[str] = set()
        for r in records:
            match, rule = match_record(r, idxs)
            matches[r["project_id"]] = (match, rule)
            if rule.startswith("AMBIGUOUS"):
                report["ambiguous_matches"].append({"project_id": r["project_id"], "title": r["title"], "reason": rule})
            elif match:
                report["matched"] += 1
                report["to_update"] += 1
                matched_ids.add(str(match["id"]))
            else:
                report["to_create"] += 1
        report["unmatched_existing"] = [p for p in existing if str(p.get("id")) not in matched_ids]
    else:
        report["backend_note"] = "No Supabase credentials present; workbook validation completed but backend reconciliation was not executed."

    if args.apply:
        if not api:
            raise SystemExit("--apply requires SUPABASE_URL and SUPABASE_SERVICE_ROLE_KEY")
        if os.getenv(WRITE_AUTHORIZATION_ENV) != WRITE_AUTHORIZATION_PHRASE:
            raise SystemExit(
                f"--apply blocked: set {WRITE_AUTHORIZATION_ENV} to the exact approved production authorization phrase."
            )
        if duplicate_ids or required_failures or report["ambiguous_matches"]:
            raise SystemExit("Apply blocked: resolve duplicate/required-field/team-size/data-link/ambiguous-match issues in the dry-run report first.")
        counts = {"updated": 0, "created": 0, "unchanged": 0}
        writes = {"project": 0, "brief": 0, "deliverables": 0, "success_criteria": 0, "roles": 0, "sources": 0}
        for r in records:
            match, _ = matches[r["project_id"]]
            action, record_writes = apply_record(api, r, match)
            counts[action] += 1
            for key_name, value in record_writes.items():
                writes[key_name] += value
        report["apply_results"] = counts
        report["write_counts"] = writes
        report["total_writes"] = sum(writes.values())

    args.report.write_text(json.dumps(report, indent=2, ensure_ascii=False), encoding="utf-8")
    print(json.dumps({k: report[k] for k in [
        "workbook_projects", "unique_project_ids", "projects_with_complete_role_responsibilities",
        "team_size_distribution", "preservation_distribution", "backend_before", "matched",
        "to_update", "to_create", "ambiguous_matches", "apply",
    ]}, indent=2))
    if args.apply:
        print(json.dumps({"apply_results": report["apply_results"], "write_counts": report["write_counts"], "total_writes": report["total_writes"]}, indent=2))
    print(f"Report: {args.report}")
    return 0 if not required_failures and not duplicate_ids else 2


if __name__ == "__main__":
    raise SystemExit(main())
