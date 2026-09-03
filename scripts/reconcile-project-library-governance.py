#!/usr/bin/env python3
"""Reconcile workbook pathway/governance metadata and stale canonical children.

Run this after scripts/import-project-library.py. Default mode is DRY RUN.
The script intentionally updates only canonical editorial/template records and never
removes run-scoped operational history.
"""
from __future__ import annotations

import argparse
import json
import os
import re
import urllib.error
import urllib.parse
import urllib.request
from dataclasses import dataclass
from pathlib import Path
from typing import Any

from openpyxl import load_workbook

PROJECT_SHEET = "03_PROJECT_LIBRARY"
PROJECT_HEADER_ROW = 4
WRITE_AUTHORIZATION_ENV = "PROJECT_LIBRARY_WRITE_AUTHORIZATION"
WRITE_AUTHORIZATION_PHRASE = "I AUTHORIZE THE PRODUCTION PROJECT LIBRARY APPLY"
SUPPORTING_SHEETS = {
    "06_DATASET_PRESERVATION": 6,
    "07_SOURCES": 3,
    "08_PROJECT_REVIEW": 3,
    "09_DIRECTOR_QUALITY_AUDIT": 1,
}


def text(value: Any) -> str:
    return "" if value is None else str(value).strip()


def norm(value: Any) -> str:
    return re.sub(r"\s+", " ", text(value)).casefold()


def yes(value: Any) -> bool:
    return norm(value).startswith("yes")


def read_table(ws, header_row: int) -> tuple[list[str], list[dict[str, Any]]]:
    headers = [text(cell.value) for cell in ws[header_row]]
    rows: list[dict[str, Any]] = []
    for values in ws.iter_rows(min_row=header_row + 1, values_only=True):
        if not any(value is not None and text(value) for value in values):
            continue
        rows.append({headers[i]: values[i] for i in range(min(len(headers), len(values))) if headers[i]})
    return headers, rows


@dataclass
class Api:
    base: str
    key: str

    def request(self, method: str, path: str, payload: Any | None = None, prefer: str | None = None) -> Any:
        body = None if payload is None else json.dumps(payload).encode()
        headers = {"apikey": self.key, "Authorization": f"Bearer {self.key}", "Accept": "application/json"}
        if body is not None:
            headers["Content-Type"] = "application/json"
        if prefer:
            headers["Prefer"] = prefer
        req = urllib.request.Request(
            self.base.rstrip("/") + "/rest/v1/" + path.lstrip("/"),
            data=body,
            method=method,
            headers=headers,
        )
        try:
            with urllib.request.urlopen(req, timeout=90) as response:
                raw = response.read().decode()
                return json.loads(raw) if raw else None
        except urllib.error.HTTPError as exc:
            detail = exc.read().decode(errors="replace")
            raise RuntimeError(f"Supabase {method} {path} failed: HTTP {exc.code}: {detail}") from exc

    def patch(self, table: str, query: str, payload: dict[str, Any]) -> None:
        self.request("PATCH", f"{table}?{query}", payload, "return=minimal")

    def delete(self, table: str, query: str) -> None:
        self.request("DELETE", f"{table}?{query}", None, "return=minimal")


def canonical_project_rows(wb) -> list[dict[str, Any]]:
    ws = wb[PROJECT_SHEET]
    _, rows = read_table(ws, PROJECT_HEADER_ROW)
    return [row for row in rows if text(row.get("Project ID"))]


def validate_workbook(wb, projects: list[dict[str, Any]]) -> list[str]:
    issues: list[str] = []
    ids = [text(row.get("Project ID")) for row in projects]
    id_set = set(ids)
    if len(projects) != 300:
        issues.append(f"PROJECT_COUNT:expected=300:actual={len(projects)}")
    if len(id_set) != len(ids):
        issues.append("DUPLICATE_PROJECT_IDS")

    for sheet, header_row in SUPPORTING_SHEETS.items():
        if sheet not in wb.sheetnames:
            issues.append(f"MISSING_SUPPORTING_SHEET:{sheet}")
            continue
        _, rows = read_table(wb[sheet], header_row)
        if sheet == "06_DATASET_PRESERVATION":
            supported: set[str] = set()
            for row in rows:
                for project_id in re.split(r"[,;\n]+", text(row.get("Project ID(s)"))):
                    if project_id.strip():
                        supported.add(project_id.strip())
            missing = sorted(id_set - supported)
            if missing:
                issues.append("PRESERVATION_MISSING_PROJECTS:" + ",".join(missing))
        else:
            support_ids = {text(row.get("Project ID")) for row in rows if text(row.get("Project ID"))}
            missing = sorted(id_set - support_ids)
            if missing:
                issues.append(f"{sheet}_MISSING_PROJECTS:" + ",".join(missing))

    source_rows = {text(row.get("Project ID")): row for row in read_table(wb["07_SOURCES"], 3)[1] if text(row.get("Project ID"))}
    review_rows = {text(row.get("Project ID")): row for row in read_table(wb["08_PROJECT_REVIEW"], 3)[1] if text(row.get("Project ID"))}
    director_rows = {text(row.get("Project ID")): row for row in read_table(wb["09_DIRECTOR_QUALITY_AUDIT"], 1)[1] if text(row.get("Project ID"))}

    for row in projects:
        pid = text(row.get("Project ID"))
        src = source_rows.get(pid)
        if src:
            if norm(src.get("Working Link")) != norm(row.get("Data Link")):
                issues.append(f"SOURCE_LINK_MISMATCH:{pid}")
            if norm(src.get("Dataset")) != norm(row.get("Dataset")):
                issues.append(f"SOURCE_DATASET_MISMATCH:{pid}")
        review = review_rows.get(pid)
        if review and not norm(review.get("Decision")).startswith("approve"):
            issues.append(f"PROJECT_REVIEW_NOT_APPROVED:{pid}:{text(review.get('Decision'))}")
        director = director_rows.get(pid)
        if director:
            expected_status = norm(row.get("Mettelo Content Quality Status"))
            if expected_status and norm(director.get("Content Status")) != expected_status:
                issues.append(f"DIRECTOR_STATUS_MISMATCH:{pid}")
            preservation = text(row.get("Preservation Class")).upper()
            if preservation and text(director.get("Preservation Class")).upper() != preservation:
                issues.append(f"DIRECTOR_PRESERVATION_MISMATCH:{pid}")
            try:
                role_count = int(director.get("Roles"))
                role_tokens = [part.strip() for part in text(row.get("Team / Roles")).split(";") if part.strip()]
                if ";" in text(row.get("Team / Roles")) and role_count != len(role_tokens):
                    issues.append(f"DIRECTOR_ROLE_COUNT_MISMATCH:{pid}:audit={role_count}:library={len(role_tokens)}")
            except (TypeError, ValueError):
                pass
    return issues


def project_metadata(row: dict[str, Any]) -> dict[str, Any]:
    return {
        "career_domain_path": text(row.get("Career / Domain Path")) or None,
        "target_role": text(row.get("Target Role")) or None,
        "path_project_number": text(row.get("Path Project #")) or None,
        "path_stage": text(row.get("Path Stage")) or None,
        "competency_focus": text(row.get("Competency Focus")) or None,
        "capability_built": text(row.get("Capability Built")) or None,
        "prerequisite_prior_project": text(row.get("Prerequisite / Prior Project")) or None,
        "path_outcome": text(row.get("Path Outcome")) or None,
        "content_quality_status": text(row.get("Mettelo Content Quality Status")) or None,
        "director_review_note": text(row.get("Director Review Note")) or None,
    }


def source_metadata(row: dict[str, Any]) -> dict[str, Any]:
    return {
        "may_redistribute": yes(row.get("Mettelo May Redistribute?")),
        "commercial_reuse": text(row.get("Commercial Reuse?")) or None,
        "attribution_required": text(row.get("Attribution Required?")) or None,
        "recommended_archive_format": text(row.get("Recommended Archive Format")) or None,
        "preservation_action": text(row.get("Preservation Action")) or None,
        "legal_review_basis": text(row.get("Legal Review Basis")) or None,
        "last_classification_review": text(row.get("Last Classification Review")) or None,
        "preservation_mode": text(row.get("Preservation Mode")) or None,
    }


def expected_child_keys(pid: str, row: dict[str, Any]) -> dict[str, set[str]]:
    def list_items(value: Any) -> list[str]:
        raw = text(value)
        if not raw:
            return []
        values = [part.strip() for part in re.split(r"\r?\n", raw) if part.strip()]
        if len(values) == 1 and ";" in raw:
            values = [part.strip() for part in raw.split(";") if part.strip()]
        return [re.sub(r"^\s*(?:[-*•]|\d{1,3}[.)])\s*", "", part).strip() for part in values if part.strip()]

    deliverables = list_items(row.get("Specific Deliverables"))
    criteria = list_items(row.get("Success Criteria"))
    # The primary importer owns role parsing. Stale-role cleanup uses its deterministic
    # key sequence and therefore only retires keys beyond the current role count.
    team_raw = text(row.get("Team / Roles"))
    if ";" in team_raw:
        role_count = len([part for part in team_raw.split(";") if part.strip()])
    elif team_raw.lower().startswith("solo project"):
        role_count = 1
    else:
        role_count = len([part for part in re.split(r"(?<=\.)\s+(?=[A-Z][^.]+?\s+(?:—|–))", team_raw) if part.strip()]) or 1
    return {
        "project_deliverables": {f"{pid}:deliverable:{i:03d}" for i in range(1, len(deliverables) + 1)},
        "project_success_criteria": {f"{pid}:criterion:{i:03d}" for i in range(1, len(criteria) + 1)},
        "project_roles": {f"{pid}:role:{i:02d}" for i in range(1, role_count + 1)},
        "project_data_sources": {f"{pid}:source:01"},
    }


def reconcile(api: Api, projects: list[dict[str, Any]], apply: bool) -> dict[str, Any]:
    report = {"projects_found": 0, "metadata_updates": 0, "source_updates": 0, "stale_children": [], "missing_backend_projects": []}
    for row in projects:
        pid = text(row.get("Project ID"))
        encoded = urllib.parse.quote(pid)
        matches = api.request("GET", f"projects?select=id,canonical_project_key,title&canonical_project_key=eq.{encoded}&limit=2") or []
        if not matches:
            title = urllib.parse.quote(text(row.get("Project Title")))
            matches = api.request("GET", f"projects?select=id,canonical_project_key,title&title=eq.{title}&limit=2") or []
        if len(matches) != 1:
            report["missing_backend_projects"].append({"project_id": pid, "matches": len(matches)})
            continue
        report["projects_found"] += 1
        project_uuid = str(matches[0]["id"])
        qid = urllib.parse.quote(project_uuid)

        brief_payload = project_metadata(row)
        source_payload = source_metadata(row)
        report["metadata_updates"] += 1
        report["source_updates"] += 1
        if apply:
            api.patch("project_problem_briefs", f"project_id=eq.{qid}", brief_payload)
            api.patch("project_data_sources", f"project_id=eq.{qid}&project_run_id=is.null&canonical_source_key=not.is.null", source_payload)

        expected = expected_child_keys(pid, row)
        child_specs = {
            "project_deliverables": ("canonical_item_key", f"project_id=eq.{qid}&project_run_id=is.null&canonical_item_key=not.is.null"),
            "project_success_criteria": ("canonical_item_key", f"project_id=eq.{qid}&canonical_item_key=not.is.null"),
            "project_roles": ("canonical_role_key", f"project_id=eq.{qid}&canonical_role_key=not.is.null"),
            "project_data_sources": ("canonical_source_key", f"project_id=eq.{qid}&project_run_id=is.null&canonical_source_key=not.is.null"),
        }
        for table, (field, query) in child_specs.items():
            rows = api.request("GET", f"{table}?select=id,{field}&{query}") or []
            for child in rows:
                key = text(child.get(field))
                if key and key not in expected[table]:
                    finding = {"project_id": pid, "table": table, "key": key, "id": str(child["id"])}
                    report["stale_children"].append(finding)
                    if apply:
                        api.delete(table, "id=eq." + urllib.parse.quote(str(child["id"])))
    return report


def main() -> int:
    parser = argparse.ArgumentParser()
    parser.add_argument("workbook", type=Path)
    parser.add_argument("--apply", action="store_true")
    parser.add_argument("--report", type=Path, default=Path("project-library-governance-report.json"))
    args = parser.parse_args()

    wb = load_workbook(args.workbook, read_only=True, data_only=True)
    projects = canonical_project_rows(wb)
    issues = validate_workbook(wb, projects)
    report: dict[str, Any] = {"workbook_projects": len(projects), "validation_issues": issues, "apply": args.apply}

    url = os.getenv("SUPABASE_URL")
    key = os.getenv("SUPABASE_SERVICE_ROLE_KEY")
    if url and key:
        if args.apply and os.getenv(WRITE_AUTHORIZATION_ENV) != WRITE_AUTHORIZATION_PHRASE:
            raise SystemExit(f"--apply blocked: set {WRITE_AUTHORIZATION_ENV} to the approved authorization phrase")
        if args.apply and issues:
            raise SystemExit("--apply blocked: supporting-sheet/project-governance validation issues must be resolved first")
        report["reconciliation"] = reconcile(Api(url, key), projects, args.apply)
    else:
        report["backend_note"] = "No Supabase credentials; workbook governance validation only."

    args.report.write_text(json.dumps(report, indent=2, ensure_ascii=False), encoding="utf-8")
    print(json.dumps(report, indent=2, ensure_ascii=False))
    return 0 if not issues else 2


if __name__ == "__main__":
    raise SystemExit(main())
